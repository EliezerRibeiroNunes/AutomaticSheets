from tkinter import filedialog, messagebox, Scrollbar, Button, Frame, Listbox, Tk, END
import fitz
import customtkinter
import openpyxl
import xlrd
import os
import json
import re
from openpyxl.styles import NamedStyle
from openpyxl.styles import Alignment
from controllers.main_controller import MainController
from controllers.db_controller import DBController

customtkinter.set_appearance_mode("System") 
customtkinter.set_default_color_theme("blue") 

class MainView:
    def __init__(self, main_controller):
        self.controller = main_controller
        self.main_controller = MainController()
        self.db_controller = DBController()
        
        with open('database.json', "r") as database:
            self.app_data = json.load(database)
        
        self.output_directory =  self.app_data['path_directory']
        self.capital_value =  self.app_data['capital_value']
        
        self.window = customtkinter.CTk()
        self.window.title("SIGAFY APP")
        self.window.geometry(f"{1100}x{580}")

        # configure grid layout (4x4)
        self.window.grid_columnconfigure(1, weight=1)
        self.window.grid_columnconfigure((2, 3), weight=0)
        self.window.grid_rowconfigure((0, 1, 2), weight=1)

        # create sidebar frame with widgets
        self.sidebar_frame = customtkinter.CTkFrame(self.window, width=140, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=4, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(4, weight=1)
        
        #self.logo_label = customtkinter.CTkLabel(self.sidebar_frame, text="SIGAFY ", font=customtkinter.CTkFont(size=30, weight="bold"))
        #self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))
        
        # Sidebar Buttons
        self.sidebar_button_1 = customtkinter.CTkButton(self.sidebar_frame, text="Selecionar Arquivos", command=self.open_file_manager,)
        self.sidebar_button_1.grid(row=1, column=0, padx=20, pady=10)
        
        self.sidebar_button_2 = customtkinter.CTkButton(self.sidebar_frame, fg_color='red', text="Excluir Todos", command=self.delete_all_files)
        self.sidebar_button_2.grid(row=2, column=0, padx=20, pady=10)
        
        self.sidebar_button_3 = customtkinter.CTkButton(self.sidebar_frame, fg_color='red', text="Excluir Arquivo", command=self.delete_selected_file)
        self.sidebar_button_3.grid(row=3, column=0, padx=20, pady=10)
        
        self.appearance_mode_label = customtkinter.CTkLabel(self.sidebar_frame, text="Tema da interface:", anchor="w")
        self.appearance_mode_label.grid(row=5, column=0, padx=20, pady=(10, 0))
        
        self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["Light", "Dark", "System"],command=self.change_appearance_mode_event)
        self.appearance_mode_optionemenu.grid(row=6, column=0, padx=20, pady=(10, 10))
        
        self.scaling_label = customtkinter.CTkLabel(self.sidebar_frame, text="Escala da interface:", anchor="w")
        self.scaling_label.grid(row=7, column=0, padx=20, pady=(10, 0))
        self.scaling_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["80%", "90%", "100%", "110%", "120%"],command=self.change_scaling_event)
        self.scaling_optionemenu.grid(row=8, column=0, padx=20, pady=(10, 20))

        #FRAMES    
        self.list_selected_files = Listbox(self.window)
        self.list_selected_files.grid(row=0, column=1, columnspan=3, padx=(20, 0), pady=(20, 0), sticky="nsew")
        
        ctk_selected_files_scrollbar = customtkinter.CTkScrollbar(self.window, command=self.list_selected_files.yview)
        ctk_selected_files_scrollbar.grid(row=0, column=4, padx=(0, 0), pady=(20, 0), sticky="ns")
        self.list_selected_files.configure(yscrollcommand=ctk_selected_files_scrollbar.set)
            
        self.textbox = customtkinter.CTkTextbox(self.window, width=250)
        self.textbox.grid(row=1, column=1, padx=(20, 0), pady=(20, 0), sticky="nsew")
        
        self.actions_frame = customtkinter.CTkFrame(self.window)
        self.actions_frame.grid(row=1, column=3, padx=(20, 0), pady=(20, 0), sticky="nsew")
        
        self.sidebar_button_3 = customtkinter.CTkButton(self.actions_frame, fg_color='green', text="Criar Planilhas", command=self.extract_data)
        self.sidebar_button_3.grid(row=1, column=1,  padx=(20, 20), pady=(20, 0))
        
        self.sidebar_button_4 = customtkinter.CTkButton(self.actions_frame, text="Configurar Diretório", command=self.set_directory)
        self.sidebar_button_4.grid(row=2, column=1,  padx=(20, 20), pady=(20, 0))
        
        self.string_input_button = customtkinter.CTkButton(self.actions_frame, text="Valor do capital", command=self.open_input_dialog_event)
        self.string_input_button.grid(row=3, column=1,  padx=(20, 20), pady=(20, 0))
         
        self.window.mainloop()
    
    def is_valid_number(input_string):
        return re.match(r'^\d+(\.\d+)?$', input_string) is not None

    def open_input_dialog_event(self):
        with open('database.json', "r") as database:
            app_data = json.load(database)
        
        formatted_value = 'R$ {:,.2f}'.format(float(app_data['capital_value'])).replace('.', 'X').replace(',', '.').replace('X', ',')
        
        dialog = customtkinter.CTkInputDialog(text=f"Valor Atual: {formatted_value}", title="Valor do capital")
        input_value = dialog.get_input()

        if input_value is not None:
    
            if input_value.isdigit():
                if len(input_value) > 11:
                    messagebox.showerror("Erro", "O valor inserido ultrapassa o limite estabelecido (11 Digitos)")
                    return
                else:
                    self.db_controller.update_database_value(float(input_value), 'capital_value')
                    messagebox.showinfo("Informativo", f"Valor atualizado para: \n{'R$ {:,.2f}'.format(float(input_value)).replace('.', 'X').replace(',', '.').replace('X', ',')}")
                    return
            else:
                messagebox.showerror("Erro", "Por favor, insira um valor válido.")
            
    def open_file_manager(self):
        
        file_paths = filedialog.askopenfilenames(
            title="Selecionar Arquivos",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls")],
            multiple=True  
         )

        if file_paths:
            for file_path in file_paths:
                self.list_selected_files.insert(END, file_path)
                
            return file_paths
        else:
            return None
        
    def delete_selected_file(self):
        selected_files = self.list_selected_files.curselection()

        if not selected_files:
            messagebox.showwarning("Aviso", "Selecione um arquivo")
            return
        else:
            for index in reversed(selected_files):
                self.list_selected_files.delete(index)
    
    def delete_all_files(self):
        self.list_selected_files.delete(0, 'end')

    def set_directory(self):    
        directory = filedialog.askdirectory(
            title="Diretório de salvamento"
            #initialdir=app_data['path_directory']
        )
        
        self.db_controller.update_database_value(directory, 'path_directory')
        
        with open('database.json', "r") as database:
            app_data = json.load(database)
            
        self.output_directory = app_data['path_directory']
        
        if  self.output_directory:
            messagebox.showinfo("Informativo", f"Diretório definido como:\n{self.output_directory}")
        else:
            messagebox.showwarning("Aviso", "Diretório não definido!.")
            
    def extract_data(self):
        selected_file_paths = self.list_selected_files.get(0, END)
        
        if not selected_file_paths:
            messagebox.showwarning("Aviso", "Nenhum arquivo selecionado!")
            return
        
        if not self.output_directory:
            messagebox.showwarning("Aviso", "Diretório para salvar os arquivos não foi definido!")
            return
        
        date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
        
        app_data = self.db_controller.get_all_data()
        
        for file_path in selected_file_paths:
            original_filename = os.path.basename(file_path)
                
            try:
                input_file = xlrd.open_workbook(file_path)
                input_sheet = input_file.sheet_by_index(0)
                
                model_workbook = openpyxl.load_workbook('ARQUIVOMODELO.xlsx')
                model_sheet = model_workbook.active
                 
                for row_num in range(1, input_sheet.nrows):
                    
                    col_name = input_sheet.cell_value(row_num, 3)
                    col_birthdate = input_sheet.cell_value(row_num, 6)
                    col_cpf = input_sheet.cell_value(row_num, 7)
                    
                    if col_cpf:
                        model_sheet.cell(row=row_num + 1, column=1, value=col_name)
                        model_sheet.cell(row=row_num + 1, column=2, value=col_cpf)
                        model_sheet.cell(row=row_num + 1, column=3, value=col_birthdate).style = date_style
                        
                        formatted_value = '{:,.2f}'.format(float(app_data['capital_value'])).replace('.', 'X').replace(',', '.').replace('X', ',')
                        
                        cell = model_sheet.cell(row=row_num + 1, column=4, value=formatted_value)
                        cell.alignment = Alignment(horizontal='right')
                                                       
                #new_filename = os.path.splitext(original_filename)[0] + '_gerado.xlsx'
                output_file_path = os.path.join(self.output_directory, original_filename)
                model_workbook.save(output_file_path)
                    
                self.textbox.insert(END, f"{original_filename} ----- CONCLUÍDO\n")
                self.textbox.update_idletasks()  
                
            except FileNotFoundError:
                self.textbox.insert(END, f"{original_filename} ----- FALHA: Arquivo não encontrado:\n")
                self.textbox.update_idletasks()
            except ValueError as e:
                self.textbox.insert(END, f"{original_filename} ----- FALHA: {str(e)}\n")
                self.textbox.update_idletasks()
            except Exception as e:
                self.textbox.insert(END, f"{original_filename} ----- FALHA: {str(e)}\n")
                self.textbox.update_idletasks()
               
        self.list_selected_files.delete(0, END)
        messagebox.showinfo("Informativo", "Processo Finalizado")

    def update_list_extracted_files(self, info_file):
        self.list_extracted_files.insert(END, info_file)
        
    def change_appearance_mode_event(self, new_appearance_mode: str):
        customtkinter.set_appearance_mode(new_appearance_mode)

    def change_scaling_event(self, new_scaling: str):
        new_scaling_float = int(new_scaling.replace("%", "")) / 100
        customtkinter.set_widget_scaling(new_scaling_float)

